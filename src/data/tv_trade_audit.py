"""TradingView trade-list audit helpers.

The replay loader only needs dated R-multiples. Deployment compliance also
needs entry/exit timestamps, duration, and trade-level P&L. This module parses
TradingView's paired entry/exit rows into complete trade records for that use.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.data.tv_trade_loader import _HeaderNotFound, _column_index, _parse_number, _parse_session_date


@dataclass(frozen=True)
class TvTradeRecord:
    trade_number: int
    entry_time: datetime
    exit_time: datetime
    net_profit: float

    @property
    def hold_seconds(self) -> float:
        return (self.exit_time - self.entry_time).total_seconds()


_MAX_REALISTIC_TRADE_PNL = 10_000.0  # abs(net_profit) cap — filters TradingView phantom open-position artifacts


def load_tv_trade_records_xlsx(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    max_abs_profit: float = _MAX_REALISTIC_TRADE_PNL,
) -> list[TvTradeRecord]:
    """Load paired TradingView entry/exit rows into complete trade records."""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet_names = [sheet_name] if sheet_name is not None else workbook.sheetnames
        last_error: ValueError | None = None
        for candidate_name in worksheet_names:
            if candidate_name not in workbook.sheetnames:
                raise ValueError(f"worksheet not found: {candidate_name}")
            worksheet = workbook[candidate_name]
            try:
                return _records_from_rows(list(worksheet.iter_rows(values_only=True)), max_abs_profit=max_abs_profit)
            except _HeaderNotFound as exc:
                last_error = exc
                if sheet_name is not None:
                    raise
        raise ValueError("no TradingView trade-list table found") from last_error
    finally:
        workbook.close()


def _records_from_rows(rows: list[tuple[Any, ...]], *, max_abs_profit: float = _MAX_REALISTIC_TRADE_PNL) -> list[TvTradeRecord]:
    header_index, headers = _find_trade_header(rows)
    trade_index = _column_index(
        headers,
        requested=None,
        candidates=("trade #", "trade number", "trade-nummer", "trade nummer"),
        label="trade #",
    )
    type_index = _column_index(headers, requested=None, candidates=("type", "typ"), label="trade type")
    time_index = _column_index(
        headers,
        requested=None,
        candidates=("date/time", "date time", "exit time", "datum und uhrzeit"),
        label="date/time",
    )
    profit_index = _column_index(
        headers,
        requested=None,
        candidates=(
            "net profit",
            "net profit usd",
            "net pnl",
            "net pnl usd",
            "g&v netto",
            "g&v netto usd",
            "netto g&v",
            "netto g&v usd",
            "netto gv",
            "netto gv usd",
        ),
        label="profit",
        banned=("cum profit", "cumulative profit", "cumulative pnl"),
    )

    grouped: dict[int, dict[str, Any]] = {}
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        raw_trade = _get_cell(row, trade_index)
        if raw_trade is None or str(raw_trade).strip() == "":
            continue
        trade_number = int(raw_trade)
        row_type = str(_get_cell(row, type_index) or "").lower()
        timestamp = _parse_datetime(_get_cell(row, time_index), row_number=row_number, column=headers[time_index])
        profit = _parse_number(_get_cell(row, profit_index), row_number=row_number, column=headers[profit_index])
        bucket = grouped.setdefault(trade_number, {})
        if "exit" in row_type or "ausstieg" in row_type or "close" in row_type:
            bucket["exit_time"] = timestamp
            bucket["net_profit"] = profit
        else:
            bucket["entry_time"] = timestamp

    records: list[TvTradeRecord] = []
    for trade_number in sorted(grouped):
        bucket = grouped[trade_number]
        if {"entry_time", "exit_time", "net_profit"} <= bucket.keys():
            pnl = float(bucket["net_profit"])
            if abs(pnl) > max_abs_profit:
                continue
            records.append(
                TvTradeRecord(
                    trade_number=trade_number,
                    entry_time=bucket["entry_time"],
                    exit_time=bucket["exit_time"],
                    net_profit=pnl,
                )
            )
    return records


def _find_trade_header(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:40]):
        headers = ["" if cell is None else str(cell).strip() for cell in row]
        try:
            _column_index(
                headers,
                requested=None,
                candidates=("trade #", "trade number", "trade-nummer", "trade nummer"),
                label="trade #",
            )
            _column_index(headers, requested=None, candidates=("type", "typ"), label="trade type")
            _column_index(
                headers,
                requested=None,
                candidates=("date/time", "date time", "exit time", "datum und uhrzeit"),
                label="date/time",
            )
            _column_index(
                headers,
                requested=None,
                candidates=(
                    "net profit",
                    "net profit usd",
                    "net pnl",
                    "net pnl usd",
                    "g&v netto",
                    "g&v netto usd",
                    "netto g&v",
                    "netto g&v usd",
                    "netto gv",
                    "netto gv usd",
                ),
                label="profit",
                banned=("cum profit", "cumulative profit", "cumulative pnl"),
            )
            return index, headers
        except ValueError:
            continue
    raise _HeaderNotFound("could not find a TradingView trade-list header")


def _parse_datetime(value: Any, *, row_number: int, column: str) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        raw = value.strip().replace("T", " ")
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
        ):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                pass
    session_date = _parse_session_date(value, row_number=row_number, column=column)
    return datetime.combine(session_date, datetime.min.time())


def _get_cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None
