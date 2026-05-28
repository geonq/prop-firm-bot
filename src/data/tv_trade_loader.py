"""Load TradingView Strategy Tester exports into replay-day inputs."""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Iterable, Sequence
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from src.strategies.replay import ReplayDay


DEFAULT_DATETIME_COLUMNS = (
    "exit time",
    "exit date/time",
    "date/time",
    "date time",
    "time",
    "date",
    "datum und uhrzeit",
    "datum/uhrzeit",
    "datum uhrzeit",
    "uhrzeit",
    "datum",
)
DEFAULT_R_MULTIPLE_COLUMNS = (
    "r_multiple",
    "r multiple",
    "r-multiple",
    "r",
)
DEFAULT_PROFIT_COLUMNS = (
    "profit",
    "profit usd",
    "net profit",
    "net profit usd",
    "net pnl",
    "net pnl usd",
    "p&l",
    "pnl",
    "g&v netto",
    "g&v netto usd",
    "gv netto",
    "gv netto usd",
    "netto g&v",
    "netto g&v usd",
    "gewinn verlust netto",
    "gewinn verlust netto usd",
)
DEFAULT_TRADE_TYPE_COLUMNS = (
    "type",
    "typ",
)


class _HeaderNotFound(ValueError):
    """Raised while scanning workbook sheets that do not contain trade rows."""


def load_tv_strategy_replay_days_xlsx(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    risk_amount: float | None = None,
    datetime_column: str | None = None,
    r_multiple_column: str | None = None,
    profit_column: str | None = None,
    include_no_trade_weekdays: bool = True,
) -> list[ReplayDay]:
    """Load a TradingView Strategy Tester XLSX into dated R-multiple days.

    If the export already contains an R-multiple column, it is used directly.
    Otherwise `risk_amount` is required and the loader derives `R = profit /
    risk_amount` from a trade profit/P&L column. Empty weekdays between the
    first and last trade are included by default so timeout math remains honest.
    """
    if risk_amount is not None and risk_amount <= 0:
        raise ValueError("risk_amount must be positive")

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet_names = [sheet_name] if sheet_name is not None else workbook.sheetnames
        last_error: ValueError | None = None
        for candidate_name in worksheet_names:
            if candidate_name not in workbook.sheetnames:
                raise ValueError(f"worksheet not found: {candidate_name}")
            worksheet = workbook[candidate_name]
            try:
                return _load_from_rows(
                    list(worksheet.iter_rows(values_only=True)),
                    risk_amount=risk_amount,
                    datetime_column=datetime_column,
                    r_multiple_column=r_multiple_column,
                    profit_column=profit_column,
                    include_no_trade_weekdays=include_no_trade_weekdays,
                )
            except _HeaderNotFound as exc:
                last_error = exc
                if sheet_name is not None:
                    raise

        raise ValueError("no TradingView trade table found") from last_error
    finally:
        workbook.close()


def _load_from_rows(
    rows: Sequence[Sequence[Any]],
    *,
    risk_amount: float | None,
    datetime_column: str | None,
    r_multiple_column: str | None,
    profit_column: str | None,
    include_no_trade_weekdays: bool,
) -> list[ReplayDay]:
    header_index, headers = _find_header_row(
        rows,
        datetime_column=datetime_column,
        r_multiple_column=r_multiple_column,
        profit_column=profit_column,
        risk_amount=risk_amount,
    )
    datetime_index = _column_index(
        headers,
        requested=datetime_column,
        candidates=DEFAULT_DATETIME_COLUMNS,
        label="date/time",
    )
    r_index = _column_index(
        headers,
        requested=r_multiple_column,
        candidates=DEFAULT_R_MULTIPLE_COLUMNS,
        label="R-multiple",
        required=False,
    )
    profit_index = None
    if r_index is None:
        if risk_amount is None:
            raise ValueError("risk_amount is required when no R-multiple column is present")
        profit_index = _column_index(
            headers,
            requested=profit_column,
            candidates=DEFAULT_PROFIT_COLUMNS,
            label="profit",
            banned=("cum profit", "cumulative profit"),
        )
    type_index = _column_index(
        headers,
        requested=None,
        candidates=DEFAULT_TRADE_TYPE_COLUMNS,
        label="trade type",
        required=False,
    )

    rows_by_date: dict[date, list[float]] = defaultdict(list)
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if _is_blank_row(row):
            continue
        if type_index is not None and not _is_exit_row(_get_cell(row, type_index)):
            continue
        session_date = _parse_session_date(_get_cell(row, datetime_index), row_number=row_number, column=headers[datetime_index])
        if r_index is not None:
            raw_r_multiple = _get_cell(row, r_index)
            if _is_blank(raw_r_multiple):
                continue
            r_multiple = _parse_number(raw_r_multiple, row_number=row_number, column=headers[r_index])
        else:
            assert profit_index is not None
            raw_profit = _get_cell(row, profit_index)
            if _is_blank(raw_profit):
                continue
            r_multiple = _parse_number(raw_profit, row_number=row_number, column=headers[profit_index]) / risk_amount

        rows_by_date[session_date].append(r_multiple)

    return _to_replay_days(rows_by_date, include_no_trade_weekdays=include_no_trade_weekdays)


def _find_header_row(
    rows: Sequence[Sequence[Any]],
    *,
    datetime_column: str | None,
    r_multiple_column: str | None,
    profit_column: str | None,
    risk_amount: float | None,
) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:40]):
        headers = ["" if cell is None else str(cell).strip() for cell in row]
        try:
            _column_index(headers, requested=datetime_column, candidates=DEFAULT_DATETIME_COLUMNS, label="date/time")
            r_index = _column_index(
                headers,
                requested=r_multiple_column,
                candidates=DEFAULT_R_MULTIPLE_COLUMNS,
                label="R-multiple",
                required=False,
            )
            if r_index is None:
                _column_index(
                    headers,
                    requested=profit_column,
                    candidates=DEFAULT_PROFIT_COLUMNS,
                    label="profit",
                    required=True,
                    banned=("cum profit", "cumulative profit"),
                )
            return index, headers
        except ValueError:
            continue

    raise _HeaderNotFound("could not find a TradingView trade table header")


def _column_index(
    headers: Sequence[str],
    *,
    requested: str | None,
    candidates: Sequence[str],
    label: str,
    required: bool = True,
    banned: Sequence[str] = (),
) -> int | None:
    normalized_headers = [_normalize_header(header) for header in headers]
    banned_headers = {_normalize_header(header) for header in banned}
    if requested is not None:
        requested_normalized = _normalize_header(requested)
        if requested_normalized in normalized_headers:
            return normalized_headers.index(requested_normalized)
        if required:
            raise ValueError(f"missing requested {label} column: {requested}")
        return None

    for candidate in candidates:
        candidate_normalized = _normalize_header(candidate)
        for index, header in enumerate(normalized_headers):
            if header == candidate_normalized and header not in banned_headers:
                return index

    if required:
        raise ValueError(f"missing {label} column")
    return None


def _normalize_header(value: str) -> str:
    return " ".join(value.strip().lower().replace("_", " ").replace("-", " ").split())


def _parse_session_date(value: Any, *, row_number: int, column: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except (TypeError, ValueError) as exc:
            raise ValueError(f"invalid date in row {row_number}, column {column}: {value!r}") from exc

    raw = str(value).strip()
    for suffix in (" UTC", " Z"):
        if raw.endswith(suffix):
            raw = raw[: -len(suffix)]
    raw = raw.replace("T", " ")
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"invalid date in row {row_number}, column {column}: {value!r}")


def _parse_number(value: Any, *, row_number: int, column: str) -> float:
    if isinstance(value, (int, float)):
        return float(value)

    raw = str(value).strip()
    negative = raw.startswith("(") and raw.endswith(")")
    cleaned = (
        raw.strip("()")
        .replace("$", "")
        .replace("USD", "")
        .replace("usd", "")
        .replace(",", "")
        .replace("%", "")
        .replace("R", "")
        .replace("r", "")
        .strip()
    )
    try:
        number = float(cleaned)
    except ValueError as exc:
        raise ValueError(f"invalid number in row {row_number}, column {column}: {value!r}") from exc
    return -number if negative else number


def _is_exit_row(value: Any) -> bool:
    normalized = _normalize_header("" if value is None else str(value))
    return "exit" in normalized or "ausstieg" in normalized or "close" in normalized


def _to_replay_days(
    rows_by_date: dict[date, list[float]],
    *,
    include_no_trade_weekdays: bool,
) -> list[ReplayDay]:
    if not rows_by_date:
        return []

    if not include_no_trade_weekdays:
        return [
            ReplayDay(session_date=session_date, r_multiples=tuple(rows_by_date[session_date]))
            for session_date in sorted(rows_by_date)
        ]

    start = min(rows_by_date)
    end = max(rows_by_date)
    replay_days: list[ReplayDay] = []
    current = start
    while current <= end:
        if current.weekday() < 5 or current in rows_by_date:
            replay_days.append(ReplayDay(session_date=current, r_multiples=tuple(rows_by_date[current])))
        current += timedelta(days=1)
    return replay_days


def _get_cell(row: Sequence[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


def _is_blank_row(row: Iterable[Any]) -> bool:
    return all(_is_blank(cell) for cell in row)


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""
