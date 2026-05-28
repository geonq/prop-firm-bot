from pathlib import Path

from openpyxl import load_workbook

from Analysis.scripts.run_tradingview_backtest import csv_to_xlsx, rename_download


def test_csv_to_xlsx_preserves_tradingview_trade_rows(tmp_path: Path) -> None:
    csv_path = tmp_path / "report.csv"
    csv_path.write_text(
        "\ufeffTrade number,Typ,Datum und Uhrzeit,Net PnL USD\n"
        "1,Long-Ausstieg,2026-05-01 09:30,125\n",
        encoding="utf-8",
    )

    xlsx_path = csv_to_xlsx(csv_path)

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Liste der Trades"]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert rows[0] == ("Trade number", "Typ", "Datum und Uhrzeit", "Net PnL USD")
    assert rows[1] == ("1", "Long-Ausstieg", "2026-05-01 09:30", "125")


def test_rename_download_avoids_overwriting_existing_file(tmp_path: Path) -> None:
    downloaded = tmp_path / "TradingView.csv"
    existing = tmp_path / "Model_A.csv"
    downloaded.write_text("new", encoding="utf-8")
    existing.write_text("old", encoding="utf-8")

    renamed = rename_download(downloaded, output_prefix="Model_A")

    assert renamed == tmp_path / "Model_A_2.csv"
    assert renamed.read_text(encoding="utf-8") == "new"
    assert existing.read_text(encoding="utf-8") == "old"
