from pathlib import Path

import pytest
from openpyxl import Workbook

from src.data.tv_trade_loader import load_tv_strategy_replay_days_xlsx


def test_load_tv_strategy_replay_days_xlsx_derives_r_from_profit_and_fills_weekdays(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "tv_export.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "List of trades"
    worksheet.append(["TradingView Strategy Tester"])
    worksheet.append(["Trade #", "Exit Time", "Profit USD", "Cum. Profit USD"])
    worksheet.append([1, "2026-01-05 10:15", "$100.00", "$100.00"])
    worksheet.append([2, "2026-01-05 12:30", "($50.00)", "$50.00"])
    worksheet.append([3, "2026-01-07 09:45", "$250.00", "$300.00"])
    workbook.save(xlsx_path)
    workbook.close()

    days = load_tv_strategy_replay_days_xlsx(xlsx_path, risk_amount=100)

    assert [day.session_date.isoformat() for day in days] == ["2026-01-05", "2026-01-06", "2026-01-07"]
    assert days[0].r_multiples == (1.0, -0.5)
    assert days[1].r_multiples == ()
    assert days[2].r_multiples == (2.5,)


def test_load_tv_strategy_replay_days_xlsx_uses_r_column_without_risk_amount(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "tv_export.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Date/Time", "R"])
    worksheet.append(["2026-01-05 10:15", "0.8R"])
    worksheet.append(["2026-01-06 12:30", "-1"])
    workbook.save(xlsx_path)
    workbook.close()

    days = load_tv_strategy_replay_days_xlsx(xlsx_path, include_no_trade_weekdays=False)

    assert [day.session_date.isoformat() for day in days] == ["2026-01-05", "2026-01-06"]
    assert days[0].r_multiples == (0.8,)
    assert days[1].r_multiples == (-1.0,)


def test_load_tv_strategy_replay_days_xlsx_searches_sheets(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "tv_export.xlsx"
    workbook = Workbook()
    workbook.active.title = "Summary"
    workbook.active.append(["not", "trades"])
    trades = workbook.create_sheet("List of trades")
    trades.append(["Exit Date/Time", "Net Profit USD"])
    trades.append(["2026-01-05 10:15", 150])
    workbook.save(xlsx_path)
    workbook.close()

    days = load_tv_strategy_replay_days_xlsx(xlsx_path, risk_amount=100)

    assert days[0].r_multiples == (1.5,)


def test_load_tv_strategy_replay_days_xlsx_requires_risk_when_only_profit_exists(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "tv_export.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Exit Time", "Profit USD"])
    worksheet.append(["2026-01-05 10:15", 100])
    workbook.save(xlsx_path)
    workbook.close()

    with pytest.raises(ValueError, match="risk_amount"):
        load_tv_strategy_replay_days_xlsx(xlsx_path)


def test_load_tv_strategy_replay_days_xlsx_reports_bad_numeric_values(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "tv_export.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Exit Time", "R"])
    worksheet.append(["2026-01-05 10:15", "nope"])
    workbook.save(xlsx_path)
    workbook.close()

    with pytest.raises(ValueError, match="row 2"):
        load_tv_strategy_replay_days_xlsx(xlsx_path)
