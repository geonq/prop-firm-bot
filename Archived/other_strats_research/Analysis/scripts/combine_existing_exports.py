"""Combinatorial test on the 40 multi-TF XLSX exports in TVExports/.

Goal: find a combination (multi-TF voting, cross-strategy gating, parallel,
confirmation overlap) that simultaneously satisfies Profile 4:
    WR >= 0.40 AND R >= 1.7 AND freq in [2, 4] / trading day.

Per-strategy single-TF baselines are also written for reference. Output CSV:
    Analysis/output/combination_matrix_<DATE>.csv
"""
from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
TV_DIR = REPO / "TVExports"
OUT = REPO / "Analysis" / "output" / "combination_matrix_2026-05-07.csv"

DATE = "2026-05-07"

STRATEGIES = [
    "P4_Compression_Breakout_v0",
    "P4_Opening_Drive_v0",
    "P4_Pullback_Continuation_v0",
    "P4_Range_Rotation_v0",
    "P4_Regime_Switch_v0",
    "P4_Sweep_Reclaim_v0",
    "P4_Trend_Day_Stack_v0",
    "P4_VWAP_Reversion_v0",
    "OH_Pullback_v1",
    "Robust_Trend_v1",
]
TFS = ["M1", "M3", "M5", "M15"]

P4_WR = (0.40, 0.50)
P4_R = (1.7, 2.3)
P4_FREQ = (2.0, 4.0)


@dataclass(frozen=True)
class Trade:
    entry: datetime
    exit: datetime
    direction: str  # "L" or "S"
    pnl: float


def parse_datetime(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str):
        return None
    text = value.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def load_trades(path: Path) -> list[Trade]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def col(*names: str) -> int:
        for n in names:
            if n in header:
                return header.index(n)
        raise ValueError(f"missing column: tried {names} in {header}")

    typ_i = col("typ", "type")
    dt_i = col("datum und uhrzeit", "date/time")
    pnl_i = col("g&v netto usd", "p&l usd", "profit usd")
    num_i = col("trade #")

    by_num: dict[str, dict] = defaultdict(dict)
    for row in rows[1:]:
        if not row or row[num_i] is None:
            continue
        tnum = str(row[num_i]).strip()
        typ = str(row[typ_i] or "").strip().lower()
        when = row[dt_i]
        pnl = row[pnl_i]
        when = parse_datetime(when)
        if when is None:
            continue
        if "einstieg" in typ:
            by_num[tnum]["entry"] = when
            by_num[tnum]["dir"] = "L" if typ.startswith("long") else "S"
        elif "ausstieg" in typ:
            by_num[tnum]["exit"] = when
            by_num[tnum]["pnl"] = parse_number(pnl)

    out = []
    for tnum, parts in by_num.items():
        if {"entry", "exit", "dir", "pnl"} <= parts.keys():
            out.append(Trade(parts["entry"], parts["exit"], parts["dir"], parts["pnl"]))
    out.sort(key=lambda t: t.entry)
    return out


def stats(trades: list[Trade], session_days: int | None = None) -> dict:
    n = len(trades)
    if n == 0:
        return {
            "trades": 0, "days": 0, "wr": 0.0, "r": 0.0, "avg_win": 0.0,
            "avg_loss": 0.0, "freq": 0.0, "net_pnl": 0.0, "profile4": False,
        }
    wins = [t.pnl for t in trades if t.pnl > 0]
    losses = [t.pnl for t in trades if t.pnl < 0]
    wr = len(wins) / n
    avg_win = sum(wins) / len(wins) if wins else 0.0
    avg_loss = sum(losses) / len(losses) if losses else 0.0
    r = (avg_win / abs(avg_loss)) if avg_loss else 0.0

    days = session_days
    if days is None:
        days = replay_weekdays(trades)
    freq = n / days
    p4 = (P4_WR[0] <= wr <= P4_WR[1]) and (P4_R[0] <= r <= P4_R[1]) and (P4_FREQ[0] <= freq <= P4_FREQ[1])
    return {
        "trades": n, "days": days, "wr": wr, "r": r, "avg_win": avg_win,
        "avg_loss": avg_loss, "freq": freq, "net_pnl": sum(t.pnl for t in trades),
        "profile4": p4,
    }


def trading_days_union(*sets: list[Trade]) -> int:
    starts = [s[0].entry.date() for s in sets if s]
    ends = [s[-1].entry.date() for s in sets if s]
    if not starts:
        return 0
    return weekdays_between(min(starts), max(ends))


def replay_weekdays(trades: list[Trade]) -> int:
    if not trades:
        return 0
    return weekdays_between(trades[0].entry.date(), trades[-1].entry.date())


def weekdays_between(start: date, end: date) -> int:
    if end < start:
        return 0
    days = 0
    cur = start
    one = timedelta(days=1)
    while cur <= end:
        if cur.weekday() < 5:
            days += 1
        cur += one
    return max(days, 1)


# --- Combination logic --------------------------------------------------------

def filter_multi_tf_vote(primary: list[Trade], confirmer: list[Trade]) -> list[Trade]:
    """Keep `primary` trade only if `confirmer` has same-day same-direction trade."""
    confirm_keys = {(t.entry.date(), t.direction) for t in confirmer}
    return [t for t in primary if (t.entry.date(), t.direction) in confirm_keys]


def filter_gated(primary: list[Trade], gate: list[Trade]) -> list[Trade]:
    """Keep `primary` trade only if `gate` has any same-day same-direction trade.

    (Loose daily-trend gate; the gate strategy provides a per-day directional bias.)
    """
    gate_keys = {(t.entry.date(), t.direction) for t in gate}
    return [t for t in primary if (t.entry.date(), t.direction) in gate_keys]


def filter_overlap(primary: list[Trade], confirmer: list[Trade], window_min: int = 15) -> list[Trade]:
    """Keep `primary` trade if `confirmer` has same-direction entry within +/- window_min."""
    out = []
    for p in primary:
        lo = p.entry - timedelta(minutes=window_min)
        hi = p.entry + timedelta(minutes=window_min)
        for c in confirmer:
            if c.direction == p.direction and lo <= c.entry <= hi:
                out.append(p); break
    return out


def parallel_union(a: list[Trade], b: list[Trade]) -> list[Trade]:
    """Concat both streams (no dedup) and re-sort. Treats them as independent books."""
    merged = list(a) + list(b)
    merged.sort(key=lambda t: t.entry)
    return merged


# --- Driver -------------------------------------------------------------------

def xlsx(strategy: str, tf: str) -> Path | None:
    p = TV_DIR / f"{strategy}_FULL_{tf}_NQ1!_{DATE}.xlsx"
    return p if p.exists() else None


def main() -> None:
    # Load every available export once.
    cache: dict[tuple[str, str], list[Trade]] = {}
    for s in STRATEGIES:
        for tf in TFS:
            p = xlsx(s, tf)
            if not p:
                continue
            try:
                cache[(s, tf)] = load_trades(p)
            except Exception as exc:
                print(f"WARN load {s} {tf}: {exc}")
                cache[(s, tf)] = []

    rows = []

    # Baselines per (strategy, TF)
    for (s, tf), trades in cache.items():
        st = stats(trades)
        rows.append({"combo_type": "baseline", "label": f"{s}|{tf}", **st})

    def add(combo_type: str, label: str, trades: list[Trade], days: int | None = None) -> None:
        st = stats(trades, session_days=days)
        rows.append({"combo_type": combo_type, "label": label, **st})

    sr_m3 = cache.get(("P4_Sweep_Reclaim_v0", "M3"), [])
    sr_m5 = cache.get(("P4_Sweep_Reclaim_v0", "M5"), [])
    sr_m15 = cache.get(("P4_Sweep_Reclaim_v0", "M15"), [])
    rt_m15 = cache.get(("Robust_Trend_v1", "M15"), [])
    tds_m5 = cache.get(("P4_Trend_Day_Stack_v0", "M5"), [])
    vwap_m3 = cache.get(("P4_VWAP_Reversion_v0", "M3"), [])
    pbc_m3 = cache.get(("P4_Pullback_Continuation_v0", "M3"), [])

    # 1a/1b: Multi-TF voting on Sweep Reclaim
    add("multi_tf_vote", "SR_M3 ∩ SR_M5 (same-day same-dir)", filter_multi_tf_vote(sr_m3, sr_m5), days=replay_weekdays(sr_m3))
    add("multi_tf_vote", "SR_M3 ∩ SR_M15 (same-day same-dir)", filter_multi_tf_vote(sr_m3, sr_m15), days=replay_weekdays(sr_m3))
    add("multi_tf_vote", "SR_M5 ∩ SR_M15 (same-day same-dir)", filter_multi_tf_vote(sr_m5, sr_m15), days=replay_weekdays(sr_m5))
    add("multi_tf_vote", "SR_M3 ∩ SR_M5 ∩ SR_M15", filter_multi_tf_vote(filter_multi_tf_vote(sr_m3, sr_m5), sr_m15), days=replay_weekdays(sr_m3))

    # 2: Cross-strategy gating
    add("gating", "SR_M3 gated by RT_M15 same-dir day", filter_gated(sr_m3, rt_m15), days=replay_weekdays(sr_m3))
    add("gating", "SR_M5 gated by RT_M15 same-dir day", filter_gated(sr_m5, rt_m15), days=replay_weekdays(sr_m5))

    # 3: Cross-strategy parallel (union, separate books)
    par = parallel_union(sr_m3, tds_m5)
    add("parallel", "SR_M3 + TDS_M5 (parallel)", par,
        days=trading_days_union(sr_m3, tds_m5))
    par2 = parallel_union(sr_m3, vwap_m3)
    add("parallel", "SR_M3 + VWAP_M3 (parallel)", par2,
        days=trading_days_union(sr_m3, vwap_m3))

    # 4: Confirmation overlap (entry within ±15 min, same direction)
    add("overlap", "SR_M3 ∩ VWAP_M3 (±15min)", filter_overlap(sr_m3, vwap_m3, 15), days=replay_weekdays(sr_m3))
    add("overlap", "SR_M3 ∩ PBC_M3 (±15min)", filter_overlap(sr_m3, pbc_m3, 15), days=replay_weekdays(sr_m3))
    sr_then_vwap = filter_overlap(sr_m3, vwap_m3, 15)
    add("overlap", "SR_M3 ∩ VWAP_M3 ∩ PBC_M3 (±15min)", filter_overlap(sr_then_vwap, pbc_m3, 15), days=replay_weekdays(sr_m3))

    # Write CSV
    OUT.parent.mkdir(parents=True, exist_ok=True)
    fields = ["combo_type", "label", "trades", "days", "wr", "r", "avg_win",
              "avg_loss", "freq", "net_pnl", "profile4"]
    with OUT.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in fields})
    print(f"\nWrote {OUT}")

    # Print combination rows + any baseline that hits Profile 4
    print("\n=== COMBINATIONS ===")
    print(f"{'type':<14} {'label':<46} {'n':>6} {'days':>5} {'WR':>6} {'R':>6} {'freq':>6} {'pnl':>10} P4")
    for r in rows:
        if r["combo_type"] == "baseline":
            continue
        print(f"{r['combo_type']:<14} {r['label']:<46} {r['trades']:>6} {r['days']:>5} "
              f"{r['wr']*100:>5.1f}% {r['r']:>6.2f} {r['freq']:>6.2f} ${r['net_pnl']:>9,.0f} "
              f"{'YES' if r['profile4'] else '   '}")

    print("\n=== BASELINE PROFILE 4 HITS ===")
    any_hit = False
    for r in rows:
        if r["combo_type"] == "baseline" and r["profile4"]:
            any_hit = True
            print(f"  {r['label']}: WR={r['wr']*100:.1f}% R={r['r']:.2f} freq={r['freq']:.2f}")
    if not any_hit:
        print("  (none)")


if __name__ == "__main__":
    main()
